#!/usr/bin/env python3
# CASLA VOLEY — generar_datos_casla.py
# Lee hoja CASLA del CASL_GamePlan_4.xlsm y genera datos_casla.js
# Uso: python generar_datos_casla.py
import os, json, re
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl"); exit(1)

BASE  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, "CASL_GamePlan_4.xlsm")
OUT   = os.path.join(BASE, "datos_casla.js")
HOJA  = "CASLA"

# ── Helpers (mismos que actualizar_gameplan.py) ───────────────────
def si(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)))
    except: return 0

def se(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)*100))
    except: return 0

def cn(v):
    if not v: return ''
    s = re.sub(r'\s+',' ', str(v).replace('\xa0',' ')).strip()
    parts = s.split()
    i = 0
    while i < len(parts) and parts[i].isdigit(): i += 1
    return ' '.join(parts[i:]).strip()

def gn(v):
    if not v: return 0
    parts = re.sub(r'\s+',' ', str(v).replace('\xa0',' ')).strip().split()
    return int(parts[0]) if parts and parts[0].isdigit() else 0

def pct(a, b): return round(a*100/b) if b > 0 else 0

def se_pct(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)*100))
    except: return 0

ORIG = {'X5':4,'V5':4,'X6':2,'V6':2,'X8':9,'V8':9,
        'X1':3,'X7':3,'XM':3,'X2':3,'XP':8,'X4':4}

POS_COLOR = {'CENTRAL':'#f97316','OPUESTO':'#818cf8','PUNTA':'#22c55e',
             'ARMADOR':'#f59e0b','LIBERO':'#06b6d4'}

# ── Leer Excel ────────────────────────────────────────────────────
def read_casla():
    wb  = load_workbook(EXCEL, read_only=True, data_only=True, keep_vba=False)
    ws  = wb[HOJA]

    # Config desde EXPORTAR
    ws_exp   = wb['EXPORTAR']
    temporada = str(ws_exp['C3'].value or '2025/2026').strip()
    equipo    = str(ws_exp['C4'].value or 'CASLA VOLEY').strip()

    print(f"Equipo: {equipo} | Temporada: {temporada}")

    rows = list(ws.iter_rows(min_row=4, max_row=85, values_only=True))

    # ── ATAQUES rows 4-17 (misma estructura que rival) ───────────
    attack = {}

    def parse_block(rows_slice, col_start):
        cur = None
        for ri, row in enumerate(rows_slice):
            if len(row) <= col_start: continue
            nc = row[col_start]
            bc = row[col_start+1] if col_start+1 < len(row) else None
            tc = row[col_start+11] if col_start+11 < len(row) else None
            if nc and gn(nc) > 0:
                cur = cn(nc); num = gn(nc)
                if cur not in attack:
                    attack[cur] = {'num':num, 'combos':{}}
                continue
            if cur and bc and str(bc).strip() not in ('','ZONA','TOTALES'):
                cod = str(bc).strip()
                tot = si(tc)
                zones = {}
                for di, zn in enumerate([1,9,2,6,8,5,7,4]):
                    ci = col_start+2+di
                    zones[zn] = si(row[ci] if ci < len(row) else None)
                attack[cur]['combos'][cod] = {'total':tot, 'zones':zones}

    parse_block(rows[0:7],  0)
    parse_block(rows[0:7],  12)
    parse_block(rows[8:14], 0)
    parse_block(rows[8:14], 12)
    parse_block(rows[0:5],  24)
    parse_block(rows[0:5],  36)
    parse_block(rows[0:5],  48)
    parse_block(rows[8:14], 24)
    parse_block(rows[8:14], 36)
    parse_block(rows[8:14], 48)

    # ── SAQUES rows 22-34 ────────────────────────────────────────
    serve = {}
    saq_rows = list(ws.iter_rows(min_row=22, max_row=34, values_only=True))

    def parse_serve_group(rows_slice, sm_start, sq_start):
        for block in range(6):
            col = block * 9
            nm_cell = rows_slice[sm_start][col] if col < len(rows_slice[sm_start]) else None
            if not nm_cell or not gn(nm_cell): continue
            nm = cn(nm_cell)
            if nm not in serve: serve[nm] = {'num':gn(nm_cell),'sm':{},'sq':{}}
            for ti, tipo in enumerate(['sm','sq']):
                dest_totals = {}
                for ri in range(3):
                    row = rows_slice[sm_start + ti*3 + ri]
                    for di, dz in enumerate([1,9,6,8,5,7]):
                        ci = col+2+di
                        val = si(row[ci] if ci < len(row) else None)
                        if val > 0:
                            dest_totals[dz] = dest_totals.get(dz,0) + val
                serve[nm][tipo] = dest_totals

    parse_serve_group(saq_rows, 0, 3)
    parse_serve_group(saq_rows[7:], 0, 3)

    # ── STATS rows 37-64 ─────────────────────────────────────────
    stat_rows = list(ws.iter_rows(min_row=37, max_row=64, values_only=True))
    STAT_COLS = {'sm':(0,1,2,3,4,5,6),'x8':(7,8,9,10,11,12,13),
                 'x5':(14,15,16,17,18,19,20),'x6':(21,22,23,24,25,26,27),
                 'x1':(28,29,30,31,32,33,34),'x7':(35,36,37,38,39,40,41),
                 'xp':(42,43,44,45,46,47,48)}

    def read_stat(row, cols):
        return {'eff': se(row[cols[1]] if cols[1] < len(row) else None),
                'tot': si(row[cols[2]] if cols[2] < len(row) else None),
                'pt':  si(row[cols[3]] if cols[3] < len(row) else None),
                'plus':si(row[cols[4]] if cols[4] < len(row) else None),
                'slash':si(row[cols[5]] if cols[5] < len(row) else None),
                'err': si(row[cols[6]] if cols[6] < len(row) else None)}

    stats = {}
    for row in stat_rows[:13]:
        nm = cn(row[0])
        if not nm: continue
        stats[nm] = {k: read_stat(row,v) for k,v in STAT_COLS.items()}
        stats[nm]['sq'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
        stats[nm]['v8'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
        stats[nm]['v5'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
        stats[nm]['v6'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
    for row in stat_rows[15:28]:
        nm = cn(row[0])
        if not nm or nm not in stats: continue
        stats[nm]['sq'] = read_stat(row, STAT_COLS['sm'])
        stats[nm]['v8'] = read_stat(row, STAT_COLS['x8'])
        stats[nm]['v5'] = read_stat(row, STAT_COLS['x5'])
        stats[nm]['v6'] = read_stat(row, STAT_COLS['x6'])

    # ── ARMADOR rows 67-86 (misma estructura que rival) ──────────
    arm_rows = list(ws.iter_rows(min_row=67, max_row=87, values_only=True))
    arm_tit_name = cn(arm_rows[0][1])
    arm_sup_name = cn(arm_rows[0][14]) if len(arm_rows[0]) > 14 else ''
    pos_labels = ['P1','P6','P5','P4','P3','P2']
    pos_data   = [1,3,5,7,9,11]

    def read_arm(col_off, eff_row_off):
        rots = []
        for i, pr in enumerate(pos_data):
            r_tot = arm_rows[pr]
            r_pts = arm_rows[pr+1]
            tot  = si(r_tot[10+col_off])
            dZ4  = si(r_tot[2+col_off]);  pZ4 = si(r_pts[2+col_off])
            dZ3  = si(r_tot[4+col_off]);  pZ3 = si(r_pts[4+col_off])
            dZ2  = si(r_tot[6+col_off]);  pZ2 = si(r_pts[6+col_off])
            dZ6  = si(r_tot[8+col_off]);  pZ6 = si(r_pts[8+col_off])
            rots.append({'pos':pos_labels[i],'total':tot,'dist':[
                {'zona':4,'tot':dZ4,'pts':pZ4,'pct':pct(dZ4,tot),'pct_p':pct(pZ4,dZ4)},
                {'zona':3,'tot':dZ3,'pts':pZ3,'pct':pct(dZ3,tot),'pct_p':pct(pZ3,dZ3)},
                {'zona':2,'tot':dZ2,'pts':pZ2,'pct':pct(dZ2,tot),'pct_p':pct(pZ2,dZ2)},
                {'zona':6,'tot':dZ6,'pts':pZ6,'pct':pct(dZ6,tot),'pct_p':pct(pZ6,dZ6)},
            ]})

        pills = []
        for i in range(6):
            er = arm_rows[eff_row_off + i]
            def safe_pct(idx):
                v = er[idx] if len(er) > idx else None
                return int(round(float(v)*100)) if v and str(v).strip() not in ('','None') else 0
            pills.append({'label':pos_labels[i],
                          'eff':safe_pct(30), 'tot':si(er[31] if len(er)>31 else None),
                          'pts':si(er[32] if len(er)>32 else None),
                          'pts_pct':safe_pct(33), 'err_pct':safe_pct(34)})

        so_row = arm_rows[eff_row_off + 6]
        tr_row = arm_rows[eff_row_off + 7]
        def pill_sotr(row, label):
            def sp(idx):
                v = row[idx] if len(row) > idx else None
                return int(round(float(v)*100)) if v and str(v).strip() not in ('','None') else 0
            return {'label':label,'eff':sp(30),'tot':si(row[31] if len(row)>31 else None),
                    'pts':si(row[32] if len(row)>32 else None),'pts_pct':sp(33),'err_pct':sp(34)}
        pills.append(pill_sotr(so_row,'SO'))
        pills.append(pill_sotr(tr_row,'TR'))

        rec_labels = ['SO','REC #','REC +','REC !','REC -','TRANS']
        rec = []
        for i in range(6):
            er = arm_rows[eff_row_off + i]
            v  = er[27] if len(er) > 27 else None
            eff_v = int(round(float(v)*100)) if v and str(v).strip() not in ('','None') else 0
            rec.append({'label':rec_labels[i],'eff':eff_v})

        d5 = [{'zona':z,'tot':0,'pts':0,'pct':0,'pct_p':0} for z in [4,3,2,8,9]]
        return {'rotaciones':rots,'pills':pills,'recepcion':rec,
                'so':{'label':'SO','total':si(so_row[31] if len(so_row)>31 else None),'dist':d5},
                'tr':{'label':'TR','total':si(tr_row[31] if len(tr_row)>31 else None),'dist':d5}}

    arm_tit = {'nombre': arm_tit_name, **read_arm(0,  2)}
    arm_sup = {'nombre': arm_sup_name, **read_arm(13, 12)}

    # ── RECEPCION rows 89-125 (misma estructura que rival) ───────
    rec_all = list(ws.iter_rows(min_row=89, max_row=126, values_only=True))
    rec_header = rec_all[1]
    rec_cols = []
    for ci, v in enumerate(rec_header):
        if v and isinstance(v, str) and any(c.isdigit() for c in str(v)[:3]):
            rec_cols.append((ci, cn(v), gn(v)))

    recepcion = {}

    def get_rec_row(row_idx, col_start):
        row = rec_all[row_idx]
        tot      = si(row[col_start+2] if col_start+2 < len(row) else None)
        eff      = se_pct(row[col_start+3] if col_start+3 < len(row) else None)
        cnt_perf = si(row[col_start+4] if col_start+4 < len(row) else None)
        cnt_pos  = si(row[col_start+5] if col_start+5 < len(row) else None)
        cnt_neg  = si(row[col_start+6] if col_start+6 < len(row) else None)
        cnt_err  = si(row[col_start+7] if col_start+7 < len(row) else None)
        return {'tot':tot,'eff':eff,
                'pos':pct(cnt_perf+cnt_pos, tot),
                'neg':pct(cnt_neg+cnt_err,  tot)}

    for col_start, name, num in rec_cols:
        flotado = {
            'desde_z1': {'total':get_rec_row(11,col_start),'p1':get_rec_row(12,col_start),
                         'p6':get_rec_row(13,col_start),'p5':get_rec_row(14,col_start)},
            'desde_z6': {'total':get_rec_row(15,col_start),'p1':get_rec_row(16,col_start),
                         'p6':get_rec_row(17,col_start),'p5':get_rec_row(18,col_start)},
            'desde_z5': {'total':get_rec_row(19,col_start),'p1':get_rec_row(20,col_start),
                         'p6':get_rec_row(21,col_start),'p5':get_rec_row(22,col_start)},
        }
        potencia = {
            'desde_z1': {'total':get_rec_row(25,col_start),'p1':get_rec_row(26,col_start),
                         'p6':get_rec_row(27,col_start),'p5':get_rec_row(28,col_start)},
            'desde_z6': {'total':get_rec_row(29,col_start),'p1':get_rec_row(30,col_start),
                         'p6':get_rec_row(31,col_start),'p5':get_rec_row(32,col_start)},
            'desde_z5': {'total':get_rec_row(33,col_start),'p1':get_rec_row(34,col_start),
                         'p6':get_rec_row(35,col_start),'p5':get_rec_row(36,col_start)},
        }
        recepcion[name] = {'num':num, 'flotado':flotado, 'potencia':potencia}

    # ── BUILD JUGADORES ───────────────────────────────────────────
    # Leer roster desde PLANTILLA_RIVAL para tener posiciones
    ws_pl = wb['PLANTILLA_RIVAL']
    roster_pos = {}
    for row in ws_pl.iter_rows(min_row=6, max_row=25, values_only=True):
        if row[2] and row[3]:
            nm = cn(row[2])
            if nm:
                roster_pos[nm] = str(row[3]).strip()

    all_names = {}
    for nm, data in attack.items():
        all_names[nm] = {'num':data['num'], 'pos':roster_pos.get(nm,'')}
    for nm, data in serve.items():
        if nm not in all_names:
            all_names[nm] = {'num':data['num'], 'pos':roster_pos.get(nm,'')}
    for nm in recepcion:
        if nm not in all_names:
            all_names[nm] = {'num':recepcion[nm]['num'], 'pos':roster_pos.get(nm,'')}

    jugadores = []
    for nm in sorted(all_names, key=lambda n: all_names[n]['num']):
        info  = all_names[nm]
        st    = stats.get(nm, {})
        atk   = attack.get(nm, {}).get('combos', {})
        saq   = serve.get(nm, {})
        pos   = info['pos']
        color = POS_COLOR.get(pos, '#64748b')

        # Ataques
        ataques = []
        for cod, a in atk.items():
            if cod in ('SM','SQ'): continue
            tot   = a.get('total', 0)
            zones = a.get('zones', {})
            key   = cod.lower()
            std   = st.get(key, {})
            eff   = std.get('eff', 0)
            pts   = std.get('pt', 0)
            if tot == 0: tot = std.get('tot', 0)
            orig  = ORIG.get(cod, 3)
            dest_sorted = sorted([(z,v) for z,v in zones.items() if v>0], key=lambda x:-x[1])
            destinos = [{'z':z,'pct':pct(v,tot)} for z,v in dest_sorted[:4]]
            ataques.append({'cod':cod,'orig':orig,'destinos':destinos,
                            'eff':eff,'tot':tot,'pts':pts,'slash':0,'err':0,
                            'video':0,'pts_pct':pct(pts,tot)})

        # Saques
        saques = []
        for cod, tipo_label in [('SM','FLOTADO'),('SQ','POTENCIA')]:
            std      = st.get(cod.lower(), {})
            tot      = std.get('tot', 0)
            eff      = std.get('eff', 0)
            pts      = std.get('pt', 0)
            saq_dest = saq.get(cod.lower(), {})
            saq_tot  = sum(saq_dest.values())
            destinos = [{'z':z,'pct':pct(v,saq_tot)}
                        for z,v in sorted(saq_dest.items(),key=lambda x:-x[1])[:4]]
            saques.append({'cod':cod,'tipo':tipo_label,'orig':6,'destinos':destinos,
                           'eff':eff,'tot':tot,'pts':pts,'plus':std.get('plus',0),
                           'slash':std.get('slash',0),'err':std.get('err',0),
                           'video':0,'pts_pct':pct(pts,tot)})

        j = {'num':info['num'],'nombre':f"{info['num']} {nm}",'pos':pos,'color':color,
             'info':{'bloqueo':'','defensa':'','pos6':'','pos5':'','pos1':'','extra':''},
             'ataques':ataques,'saques':saques,'video':0,
             'recepcion':recepcion.get(nm, {})}
        jugadores.append(j)

    print(f"✓ {len(jugadores)} jugadores de CASLA:")
    for j in jugadores:
        na = len(j['ataques'])
        ns = sum(s['tot'] for s in j['saques'])
        print(f"  #{j['num']} {cn(j['nombre'])} ({j['pos']}) — {na} combos, {ns} saques")

    return {
        'equipo':   equipo,
        'temporada': temporada,
        'generado': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'jugadores': jugadores,
        'armador':  {'titular':arm_tit, 'suplente':arm_sup},
    }

# ── Generar datos_casla.js ────────────────────────────────────────
def build(data):
    jug_json  = json.dumps(data['jugadores'],  ensure_ascii=False, indent=2)
    arm_json  = json.dumps(data['armador'],    ensure_ascii=False, indent=2)

    js = f"""// datos_casla.js — generado automáticamente por generar_datos_casla.py
// {data['generado']} — {data['equipo']} — {data['temporada']}
// NO EDITAR MANUALMENTE

const CASLA_EQUIPO = "{data['equipo']}";
const CASLA_TEMPORADA = "{data['temporada']}";
const CASLA_GENERADO = "{data['generado']}";

const CASLA_JUGADORES = {jug_json};

const CASLA_ARMADOR = {arm_json};
"""
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(js)
    print(f"\n✓ datos_casla.js generado ({len(js)//1024}KB) — subir a GitHub")

if __name__ == '__main__':
    print("="*50)
    print("CASLA VOLEY — Generar datos_casla.js")
    print("="*50)
    if not os.path.exists(EXCEL):
        print(f"ERROR: No se encontró {EXCEL}"); exit(1)
    data = read_casla()
    build(data)
