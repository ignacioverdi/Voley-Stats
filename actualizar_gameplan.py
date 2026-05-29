#!/usr/bin/env python3
# CASLA VOLEY — actualizar_gameplan.py
# Uso: correr.bat
import os, json, re
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl"); exit(1)

BASE  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, "CASL_GamePlan_4.xlsm")
TMPL  = os.path.join(BASE, "game_plan_template.html")

def si(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)))
    except: return 0
def se(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)*100))
    except: return 0
def cn(v):
    if not v: return ''
    s = re.sub(r'\s+',' ',str(v).replace('\xa0',' ')).strip()
    parts = s.split()
    i = 0
    while i < len(parts) and parts[i].isdigit(): i += 1
    return ' '.join(parts[i:]).strip()
def gn(v):
    if not v: return 0
    parts = re.sub(r'\s+',' ',str(v).replace('\xa0',' ')).strip().split()
    return int(parts[0]) if parts and parts[0].isdigit() else 0
def pct(a,b): return round(a*100/b) if b>0 else 0

# Fixed orig zone per combo (where attacker hits from)
ORIG = {'X5':4,'V5':4,'X6':2,'V6':2,'X8':9,'V8':9,
        'X1':3,'X7':3,'XM':3,'X2':3,'XP':8,'X4':4}

POS_COLOR = {'CENTRAL':'#f97316','OPUESTO':'#818cf8','PUNTA':'#22c55e',
             'ARMADOR':'#f59e0b','LIBERO':'#06b6d4'}

def read_excel():
    wb = load_workbook(EXCEL, read_only=True, data_only=True, keep_vba=False)
    ws_exp = wb['EXPORTAR']
    rival  = str(ws_exp['C3'].value or '').strip()
    torneo = str(ws_exp['C4'].value or '').strip()
    fecha  = ws_exp['C5'].value or 0
    ws = wb[rival]
    ws_vid = None
    for name in wb.sheetnames:
        if 'VIDEOS' in name.upper() and rival.upper().replace(' ','') in name.upper().replace(' ',''):
            ws_vid = wb[name]; break
    print(f"Rival: {rival} | Torneo: {torneo} | Fecha: {fecha}")
    rows = list(ws.iter_rows(min_row=4, max_row=85, values_only=True))

    # ── ATTACKS ─────────────────────────────────────────────────
    # 5 blocks of 12 cols: cols 0-11, 12-23, 24-35, 36-47, 48-59
    # Block layout: col+0=name(header), col+1=combo, col+2..9=zones(Z1,Z9,Z2,Z6,Z8,Z5,Z7,Z4), col+11=total
    # OPUESTOS rows 1-7 (row4-10): blocks 0,1
    # CENTRALES rows 1-5 (row4-8):  blocks 2,3,4
    # PUNTAS rows 9-14 (row12-17):  blocks 0,1,2,3,4
    attack = {}

    def parse_block(rows_slice, col_start):
        cur = None
        for ri, row in enumerate(rows_slice):
            if len(row) <= col_start: continue
            nc = row[col_start]
            bc = row[col_start+1] if col_start+1<len(row) else None
            tc = row[col_start+11] if col_start+11<len(row) else None
            # Player name row: has number prefix
            if nc and gn(nc) > 0:
                cur = cn(nc)
                num = gn(nc)
                if cur not in attack:
                    attack[cur] = {'num':num, 'combos':{}}
                continue
            # Combo row
            if cur and bc and str(bc).strip() not in ('','ZONA','TOTALES'):
                cod = str(bc).strip()
                tot = si(tc)
                zones = {}
                for di, zn in enumerate([1,9,2,6,8,5,7,4]):
                    ci = col_start+2+di
                    zones[zn] = si(row[ci] if ci<len(row) else None)
                attack[cur]['combos'][cod] = {'total':tot, 'zones':zones}

    # OPUESTOS + PUNTAS first two blocks (rows 1-14)
    parse_block(rows[0:7],  0)   # Opuesto left, rows 4-10
    parse_block(rows[0:7],  12)  # Opuesto right, rows 4-10
    parse_block(rows[8:14], 0)   # Punta left, rows 12-17
    parse_block(rows[8:14], 12)  # Punta right, rows 12-17
    # CENTRALES + PUNTAS blocks 2,3,4
    parse_block(rows[0:5],  24)  # Central, rows 4-8
    parse_block(rows[0:5],  36)  # Central, rows 4-8
    parse_block(rows[0:5],  48)  # Central, rows 4-8
    parse_block(rows[8:14], 24)  # Punta, rows 12-17
    parse_block(rows[8:14], 36)  # Punta, rows 12-17
    parse_block(rows[8:14], 48)  # Punta, rows 12-17

    # ── SAQUES rows 22-34 (index 18-30) ─────────────────────────
    # 6 players, 9 cols each: col+0=name, col+1=zone, col+2..7=dests(Z1,Z9,Z6,Z8,Z5,Z7), col+8=total
    # SM: rows 22-24 (index 18-20), SQ: rows 25-27 (index 21-23)
    # Group 2: rows 29-34 (index 25-30)
    serve = {}
    saq_rows = list(ws.iter_rows(min_row=22, max_row=34, values_only=True))

    def parse_serve_group(rows_slice, sm_start, sq_start):
        for block in range(6):
            col = block * 9
            nm_cell = rows_slice[sm_start][col] if col < len(rows_slice[sm_start]) else None
            if not nm_cell or not gn(nm_cell): continue
            nm = cn(nm_cell)
            if nm not in serve: serve[nm] = {'num':gn(nm_cell),'sm':{},'sq':{}}
            for ti, tipo in enumerate(['sm','sq']):
                dest_totals = {}
                for ri in range(3):
                    row = rows_slice[sm_start + ti*3 + ri]
                    for di, dz in enumerate([1,9,6,8,5,7]):
                        ci = col+2+di
                        val = si(row[ci] if ci<len(row) else None)
                        if val > 0:
                            dest_totals[dz] = dest_totals.get(dz,0) + val
                serve[nm][tipo] = dest_totals

    parse_serve_group(saq_rows, 0, 3)   # rows 22-27 (SM=0-2, SQ=3-5)
    parse_serve_group(saq_rows[7:], 0, 3) # rows 29-34

    # ── STATS rows 37-64 (index 33-60 from row 4) ───────────────
    # Header row 36: SM,EFF%,TOT,#,+,/,=, X8,EFF%,TOT,#,#%,/,=, X5,EFF%,TOT,#,#%,/,=,
    #                X6,EFF%,TOT,#,#%,/,=, X1,EFF%,TOT,#,#%,/,=, X7,EFF%,TOT,#,#%,/,=,
    #                XP,EFF%,TOT,#,#%,/,=
    stat_rows = list(ws.iter_rows(min_row=37, max_row=64, values_only=True))
    STAT_COLS = {'sm':(0,1,2,3,4,5,6),'x8':(7,8,9,10,11,12,13),
                 'x5':(14,15,16,17,18,19,20),'x6':(21,22,23,24,25,26,27),
                 'x1':(28,29,30,31,32,33,34),'x7':(35,36,37,38,39,40,41),
                 'xp':(42,43,44,45,46,47,48)}
    def read_stat(row, cols):
        return {'eff':se(row[cols[1]] if cols[1]<len(row) else None),
                'tot':si(row[cols[2]] if cols[2]<len(row) else None),
                'pt': si(row[cols[3]] if cols[3]<len(row) else None),
                'plus':si(row[cols[4]] if cols[4]<len(row) else None),
                'slash':si(row[cols[5]] if cols[5]<len(row) else None),
                'err':si(row[cols[6]] if cols[6]<len(row) else None)}
    stats = {}
    for row in stat_rows[:13]:  # SM section rows 37-49
        nm = cn(row[0])
        if not nm or 'Municipio' in nm: continue
        stats[nm] = {k:read_stat(row,v) for k,v in STAT_COLS.items()}
        stats[nm]['sq'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
        stats[nm]['v8'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
        stats[nm]['v5'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
        stats[nm]['v6'] = {'eff':0,'tot':0,'pt':0,'plus':0,'slash':0,'err':0}
    for row in stat_rows[15:28]:  # SQ section rows 52-64
        nm = cn(row[0])
        if not nm or 'Municipio' in nm or nm not in stats: continue
        stats[nm]['sq']  = read_stat(row, STAT_COLS['sm'])
        stats[nm]['v8']  = read_stat(row, STAT_COLS['x8'])
        stats[nm]['v5']  = read_stat(row, STAT_COLS['x5'])
        stats[nm]['v6']  = read_stat(row, STAT_COLS['x6'])

    # ── ARMADOR rows 67-79 ────────────────────────────────────────
    # Structure (col_off=0 for titular, 13 for suplente):
    #   Row 67 (idx0):  headers
    #   Rows 68-79 (idx1-12): P1,_,P6,_,P5,_,P4,_,P3,_,P2,_ (tot row + pts row pairs)
    #     col+0=name, col+1=pos, col+2=#P4, col+3=%P4, col+4=#P3, col+5=%P3
    #     col+6=#P2, col+7=%P2, col+8=#P6, col+9=%P6, col+10=TOT
    #   Cols 26-34 (offset from 0): EFF/recepcion + P1-P6 eff + SO/TR
    #     Row 68/78: header
    #     Rows 69-74 / 79-84: recepcion labels (col26) + eff (col27)
    #                         P-label (col29) + eff (col30) + tot (col31) + pts (col32)
    #                                         + pts_pct (col33) + err_pct (col34)
    #     Row 75/85: SO overall: col29='SD', col30=eff, col31=tot, col32=pts, col33=pts_pct, col34=err_pct
    #     Row 76/86: TR overall: col29='TR', col30=eff, col31=tot, col32=pts, col33=pts_pct, col34=err_pct
    arm_rows = list(ws.iter_rows(min_row=67, max_row=87, values_only=True))
    arm_tit_name = cn(arm_rows[0][1])
    arm_sup_name = cn(arm_rows[0][14]) if len(arm_rows[0])>14 else ''
    pos_labels = ['P1','P6','P5','P4','P3','P2']
    pos_data   = [1,3,5,7,9,11]  # 0-indexed row offsets for each position's tot row

    def read_arm(col_off, eff_row_off):
        # ── Rotaciones (canvas data) ──────────────────────────────
        rots=[]
        for i,pr in enumerate(pos_data):
            r_tot = arm_rows[pr]
            r_pts = arm_rows[pr+1]
            tot = si(r_tot[10+col_off])
            dZ4=si(r_tot[2+col_off]); pZ4=si(r_pts[2+col_off])
            dZ3=si(r_tot[4+col_off]); pZ3=si(r_pts[4+col_off])
            dZ2=si(r_tot[6+col_off]); pZ2=si(r_pts[6+col_off])
            dZ6=si(r_tot[8+col_off]); pZ6=si(r_pts[8+col_off])
            rots.append({'pos':pos_labels[i],'total':tot,'dist':[
                {'zona':4,'tot':dZ4,'pts':pZ4,'pct':pct(dZ4,tot),'pct_p':pct(pZ4,dZ4)},
                {'zona':3,'tot':dZ3,'pts':pZ3,'pct':pct(dZ3,tot),'pct_p':pct(pZ3,dZ3)},
                {'zona':2,'tot':dZ2,'pts':pZ2,'pct':pct(dZ2,tot),'pct_p':pct(pZ2,dZ2)},
                {'zona':6,'tot':dZ6,'pts':pZ6,'pct':pct(dZ6,tot),'pct_p':pct(pZ6,dZ6)},
            ]})

        # ── Pills P1-P6 from cols 29-34 ───────────────────────────
        # eff_row_off: for titular=2 (row69-74), for suplente=12 (row79-84)
        pills=[]
        for i in range(6):
            er = arm_rows[eff_row_off + i]  # row for this position
            eff_raw = er[30] if len(er)>30 else None
            eff_val  = int(round(float(eff_raw)*100)) if eff_raw and str(eff_raw).strip() not in ('','None') else 0
            tot_val  = si(er[31]) if len(er)>31 else 0
            pts_val  = si(er[32]) if len(er)>32 else 0
            ptp_raw  = er[33] if len(er)>33 else None
            ptp_val  = int(round(float(ptp_raw)*100)) if ptp_raw and str(ptp_raw).strip() not in ('','None') else 0
            erp_raw  = er[34] if len(er)>34 else None
            erp_val  = int(round(float(erp_raw)*100)) if erp_raw and str(erp_raw).strip() not in ('','None') else 0
            pills.append({'label':pos_labels[i],'eff':eff_val,'tot':tot_val,
                         'pts':pts_val,'pts_pct':ptp_val,'err_pct':erp_val})

        # ── Pills SO/TR from rows 75/76 (titular) or 85/86 (suplente) ──
        so_row = arm_rows[eff_row_off + 6]  # SD row
        tr_row = arm_rows[eff_row_off + 7]  # TR row
        def pill_sotr(row, label):
            eff_r = row[30] if len(row)>30 else None
            eff_v = int(round(float(eff_r)*100)) if eff_r and str(eff_r).strip() not in ('','None') else 0
            tot_v = si(row[31]) if len(row)>31 else 0
            pts_v = si(row[32]) if len(row)>32 else 0
            ptp_r = row[33] if len(row)>33 else None
            ptp_v = int(round(float(ptp_r)*100)) if ptp_r and str(ptp_r).strip() not in ('','None') else 0
            erp_r = row[34] if len(row)>34 else None
            erp_v = int(round(float(erp_r)*100)) if erp_r and str(erp_r).strip() not in ('','None') else 0
            return {'label':label,'eff':eff_v,'tot':tot_v,'pts':pts_v,'pts_pct':ptp_v,'err_pct':erp_v}
        pills.append(pill_sotr(so_row,'SO'))
        pills.append(pill_sotr(tr_row,'TR'))

        # ── Recepcion from cols 26-27 ─────────────────────────────
        rec_labels = ['SO','REC #','REC +','REC !','REC -','TRANS']
        rec=[]
        for i in range(6):
            er = arm_rows[eff_row_off + i]
            eff_raw = er[27] if len(er)>27 else None
            eff_val = int(round(float(eff_raw)*100)) if eff_raw and str(eff_raw).strip() not in ('','None') else 0
            rec.append({'label':rec_labels[i],'eff':eff_val})

        # ── SO/TR distribution (dist by zona 4,3,2,8,9) ──────────
        # Not in standard Excel layout — leave as empty for now
        d5=[{'zona':z,'tot':0,'pts':0,'pct':0,'pct_p':0} for z in [4,3,2,8,9]]
        return {'rotaciones':rots,'pills':pills,'recepcion':rec,
                'so':{'label':'SO','total':so_row[31] if len(so_row)>31 else 0,'dist':d5},
                'tr':{'label':'TR','total':tr_row[31] if len(tr_row)>31 else 0,'dist':d5}}

    # eff_row_off: titular rows 69-76 = indices 2-9, suplente rows 79-86 = indices 12-19
    arm_tit = {'nombre':arm_tit_name, **read_arm(0,  2)}
    arm_sup = {'nombre':arm_sup_name, **read_arm(13, 12)}


    # ── RECEPCION RIVAL rows 89-125 ──────────────────────────────
    # 6 players x 8 cols. Flotado: rows 99-111, Potencia: rows 113-125
    # Per player block: col+0=name, col+1=tipo, col+2=TOT, col+3=EFF(dec),
    #   col+4=#, col+5=+, col+6=/, col+7==
    # Row mapping:
    #   Flotado POS1: row100, sub: 101(→P1), 102(→P6), 103(→P5)
    #   Flotado POS6: row104, sub: 105(→P1), 106(→P6), 107(→P5)
    #   Flotado POS5: row108, sub: 109(→P1), 110(→P6), 111(→P5)
    #   Potencia POS1: row114, sub: 115(→P1), 116(→P6), 117(→P5)
    #   Potencia POS6: row118, sub: 119(→P1), 120(→P6), 121(→P5)
    #   Potencia POS5: row122, sub: 123(→P1), 124(→P6), 125(→P5)

    rec_all = list(ws.iter_rows(min_row=89, max_row=126, values_only=True))

    def se_pct(v):
        if v is None or str(v).strip() in ('','----'): return 0
        try: return int(round(float(v)*100))
        except: return 0

    def neg_pct(v1, v2):
        """combine slash+error into neg%"""
        a = 0
        for v in [v1, v2]:
            if v and str(v).strip() not in ('','----'):
                try: a += int(round(float(v)*100))
                except: pass
        return a

    # Find player columns from row 90 (index 1)
    rec_header = rec_all[1]
    rec_cols = []
    for ci, v in enumerate(rec_header):
        if v and isinstance(v, str) and any(c.isdigit() for c in str(v)[:3]):
            rec_cols.append((ci, cn(v), gn(v)))

    recepcion = {}

    def get_rec_row(row_idx, col_start):
        """Get {tot, eff, pos%, neg%} from a row at col_start.
        col+4=# col+5=+ col+6=/ col+7== are COUNTS, convert to %"""
        row = rec_all[row_idx]
        tot = si(row[col_start+2] if col_start+2 < len(row) else None)
        eff = se_pct(row[col_start+3] if col_start+3 < len(row) else None)
        # pos% = (# + +) / tot * 100
        cnt_perf = si(row[col_start+4] if col_start+4 < len(row) else None)
        cnt_pos  = si(row[col_start+5] if col_start+5 < len(row) else None)
        cnt_neg  = si(row[col_start+6] if col_start+6 < len(row) else None)
        cnt_err  = si(row[col_start+7] if col_start+7 < len(row) else None)
        pos_pct  = pct(cnt_perf + cnt_pos, tot)
        neg_pct_ = pct(cnt_neg  + cnt_err, tot)
        return {'tot': tot, 'eff': eff, 'pos': pos_pct, 'neg': neg_pct_}

    for col_start, name, num in rec_cols:
        # Flotado (SM): rows 100-111 = indices 11-22
        flotado = {
            'desde_z1': {
                'total': get_rec_row(11, col_start),  # row 100
                'p1':    get_rec_row(12, col_start),  # row 101
                'p6':    get_rec_row(13, col_start),  # row 102
                'p5':    get_rec_row(14, col_start),  # row 103
            },
            'desde_z6': {
                'total': get_rec_row(15, col_start),  # row 104
                'p1':    get_rec_row(16, col_start),  # row 105
                'p6':    get_rec_row(17, col_start),  # row 106
                'p5':    get_rec_row(18, col_start),  # row 107
            },
            'desde_z5': {
                'total': get_rec_row(19, col_start),  # row 108
                'p1':    get_rec_row(20, col_start),  # row 109
                'p6':    get_rec_row(21, col_start),  # row 110
                'p5':    get_rec_row(22, col_start),  # row 111
            },
        }
        # Potencia (SQ): rows 114-125 = indices 25-36
        potencia = {
            'desde_z1': {
                'total': get_rec_row(25, col_start),  # row 114
                'p1':    get_rec_row(26, col_start),  # row 115
                'p6':    get_rec_row(27, col_start),  # row 116
                'p5':    get_rec_row(28, col_start),  # row 117
            },
            'desde_z6': {
                'total': get_rec_row(29, col_start),  # row 118
                'p1':    get_rec_row(30, col_start),  # row 119
                'p6':    get_rec_row(31, col_start),  # row 120
                'p5':    get_rec_row(32, col_start),  # row 121
            },
            'desde_z5': {
                'total': get_rec_row(33, col_start),  # row 122
                'p1':    get_rec_row(34, col_start),  # row 123
                'p6':    get_rec_row(35, col_start),  # row 124
                'p5':    get_rec_row(36, col_start),  # row 125
            },
        }
        recepcion[name] = {'num': num, 'flotado': flotado, 'potencia': potencia}

    # ── VIDEOS ────────────────────────────────────────────────────
    videos = {}
    if ws_vid:
        cur_vp = None
        for row in ws_vid.iter_rows(min_row=5, max_row=400, values_only=True):
            a=row[0]; c=row[2] if len(row)>2 else None
            d=row[3] if len(row)>3 else None; f=row[5] if len(row)>5 else None
            if a and c and not str(c).startswith('=') and not isinstance(c,(int,float)):
                cur_vp = cn(a)
                if cur_vp not in videos:
                    videos[cur_vp] = {'pos':str(c).strip(),'num':gn(a),'combos':{}}
            if cur_vp and d and not str(d).startswith('='):
                cod = str(d).strip()
                url = str(f).strip() if f and not str(f).startswith('=') else ''
                videos[cur_vp]['combos'][cod] = url

    # ── BUILD JUGADORES ───────────────────────────────────────────
    all_names = {}
    for nm,data in attack.items():
        all_names[nm] = {'num':data['num'],'pos':videos.get(nm,{}).get('pos','')}
    for nm,data in serve.items():
        if nm not in all_names:
            all_names[nm] = {'num':data['num'],'pos':videos.get(nm,{}).get('pos','')}
    for nm,data in videos.items():
        if nm not in all_names:
            all_names[nm] = {'num':data['num'],'pos':data['pos']}
        else:
            if not all_names[nm]['pos']: all_names[nm]['pos'] = data['pos']
            if not all_names[nm]['num']: all_names[nm]['num'] = data['num']

    jugadores = []
    for nm in sorted(all_names, key=lambda n: all_names[n]['num']):
        info = all_names[nm]
        st   = stats.get(nm,{})
        atk  = attack.get(nm,{}).get('combos',{})
        saq  = serve.get(nm,{})
        vid  = videos.get(nm,{}).get('combos',{})
        pos  = info['pos']
        color = POS_COLOR.get(pos,'#64748b')

        # Build ataques — use EXACT structure from working DBANF
        all_cods = list(atk.keys())
        for cod in vid:
            if cod not in ('SM','SQ') and cod not in all_cods:
                all_cods.append(cod)

        ataques = []
        for cod in all_cods:
            if cod in ('SM','SQ'): continue
            a    = atk.get(cod,{})
            tot  = a.get('total',0)
            zones= a.get('zones',{})
            key  = cod.lower()
            std  = st.get(key,{})
            eff  = std.get('eff',0)
            pts  = std.get('pt',0)
            if tot == 0: tot = std.get('tot',0)
            url  = vid.get(cod,'')
            orig = ORIG.get(cod,3)
            # Build destinos: top zones by count, pct relative to total
            dest_sorted = sorted(
                [(z,v) for z,v in zones.items() if v>0],
                key=lambda x: -x[1]
            )
            destinos = [{'z':z,'pct':pct(v,tot)} for z,v in dest_sorted[:4]]
            ataques.append({
                'cod':cod,'orig':orig,'destinos':destinos,
                'eff':eff,'tot':tot,'pts':pts,
                'slash':0,'err':0,
                'video': url if url else 0,  # 0 when no video, not null
                'pts_pct':pct(pts,tot)
            })

        # Build saques
        saques = []
        for cod,tipo_label in [('SM','FLOTADO'),('SQ','POTENCIA')]:
            std  = st.get(cod.lower(),{})
            tot  = std.get('tot',0)
            eff  = std.get('eff',0)
            pts  = std.get('pt',0)
            plus = std.get('plus',0)
            slash= std.get('slash',0)
            err  = std.get('err',0)
            url  = vid.get(cod,'')
            saq_dest = saq.get(cod.lower(),{})
            saq_tot  = sum(saq_dest.values())
            destinos = [{'z':z,'pct':pct(v,saq_tot)}
                        for z,v in sorted(saq_dest.items(),key=lambda x:-x[1])[:4]]
            saques.append({
                'cod':cod,'tipo':tipo_label,'orig':6,'destinos':destinos,
                'eff':eff,'tot':tot,'pts':pts,'plus':plus,'slash':slash,'err':err,
                'video': url if url else 0,
                'pts_pct':pct(pts,tot)
            })

        jugadores.append({
            'num':info['num'],
            'nombre':f"{info['num']} {nm}",
            'pos':pos,'color':color,
            'info':{'bloqueo':'','defensa':'','pos6':'','pos5':'','pos1':'','extra':''},
            'ataques':ataques,'saques':saques,'video':0
        })

    print(f"✓ {len(jugadores)} jugadores:")
    for j in jugadores:
        na = len(j['ataques'])
        nd = sum(len(a['destinos']) for a in j['ataques'])
        ns = sum(s['tot'] for s in j['saques'])
        nv = sum(1 for a in j['ataques']+j['saques'] if a.get('video') and a['video']!=0)
        print(f"  #{j['num']} {cn(j['nombre'])} ({j['pos']}) — {na} combos, {nd} destinos, {ns} saques, {nv} videos")

    # Add recepcion data to each jugador
    for j in jugadores:
        nm = re.sub(r'^\d+\s*','',j['nombre'])
        j['recepcion'] = recepcion.get(nm, {})

    return {'rival':rival,'torneo':torneo,'fecha':fecha,
            'generado':datetime.now().strftime('%d/%m/%Y %H:%M'),
            'jugadores':jugadores,
            'armador':{'titular':arm_tit,'suplente':arm_sup}}

def build(data):
    with open(TMPL,'rb') as f:
        html = f.read().decode('utf-8',errors='replace')
    rival = data['rival']

    # Inject ARMADOR_DATA
    arm_json = json.dumps({'titular':data['armador']['titular'],
                           'suplente':data['armador']['suplente']},ensure_ascii=False)
    html = re.sub(
        r'const ARMADOR_DATA=/\*__ARM_START__\*/[\s\S]*?/\*__ARM_END__\*/;',
        f'const ARMADOR_DATA=/*__ARM_START__*/{arm_json}/*__ARM_END__*/;',
        html, count=1)

    # Inject JUGADORES
    jug_json = json.dumps(data['jugadores'],ensure_ascii=False)
    html = re.sub(
        r'const JUGADORES\s*=\s*\[[\s\S]*?\];',
        f'const JUGADORES={jug_json};',
        html, count=1)

    # Text replacements (safe — won't break JS vars like RIVAL_H)
    html = html.replace('>RIVAL<',f'>{rival.upper()}<')
    html = html.replace('"RIVAL"',f'"{rival.upper()}"')
    html = html.replace("'RIVAL'",f"'{rival.upper()}'")
    html = html.replace('TORNEO_PLACEHOLDER',data['torneo'])
    html = html.replace('FECHA_PLACEHOLDER',str(data['fecha']))

    out = f"game_plan_{rival.lower().replace(' ','_')}.html"
    with open(os.path.join(BASE,out),'w',encoding='utf-8') as f:
        f.write(html)
    print(f"\n✓ {out} generado — subir a GitHub")
    return out

if __name__=='__main__':
    print("="*50)
    print("CASLA VOLEY — Game Plan Generator")
    print("="*50)
    if not os.path.exists(EXCEL): print(f"ERROR: {EXCEL}"); exit(1)
    if not os.path.exists(TMPL):  print(f"ERROR: {TMPL}");  exit(1)
    data = read_excel()
    build(data)
