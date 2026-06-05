#!/usr/bin/env python3
"""
build_videos.py — Convierte el Excel en videos.js + proximo_rival.js

Uso:
    python3 build_videos.py videos_casla.xlsx
    python3 build_videos.py videos_nafels.xlsx

Genera:
  - videos.js         (todos los videos cargados)
  - proximo_rival.js  (el próximo rival según el FIXTURE y la fecha de hoy)

Subir ambos al repo junto al game_plan.html y el index.html.
"""
import sys, json, datetime
from openpyxl import load_workbook

def slugify(name):
    return str(name).strip().lower().replace(' ', '_')

def build_videos(wb):
    videos = {}; count = 0
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() == 'FIXTURE': continue
        ws = wb[sheet_name]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) < 6: continue
            equipo, num, jugador, tipo, combo, url = r[0], r[1], r[2], r[3], r[4], r[5]
            if not url or not str(url).strip(): continue
            if not equipo or num is None or not combo: continue
            slug = slugify(equipo)
            num = str(int(num)) if isinstance(num,(int,float)) else str(num).strip()
            tipo = str(tipo).strip().lower()
            combo = str(combo).strip().upper()
            videos.setdefault(slug, {}).setdefault(num, {}).setdefault(tipo, {})[combo] = str(url).strip()
            count += 1
    with open('videos.js','w',encoding='utf-8') as f:
        f.write('window.VIDEOS_DATA = ' + json.dumps(videos, ensure_ascii=False) + ';\n')
    print(f"  videos.js: {count} videos")
    return count

def build_fixture(wb):
    if 'FIXTURE' not in [s.upper() for s in wb.sheetnames]:
        print("  (sin solapa FIXTURE, no genero proximo_rival.js)")
        return None
    # find sheet (case-insensitive)
    fx_name = next(s for s in wb.sheetnames if s.upper()=='FIXTURE')
    ws = wb[fx_name]
    partidos = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 2: continue
        fecha, rival = r[0], r[1]
        if not fecha or not rival: continue
        # parse fecha
        if isinstance(fecha, datetime.datetime):
            fdate = fecha.date()
        elif isinstance(fecha, datetime.date):
            fdate = fecha
        else:
            try: fdate = datetime.datetime.strptime(str(fecha).strip()[:10], '%Y-%m-%d').date()
            except: continue
        cond = r[2] if len(r)>2 and r[2] else ''
        partidos.append({'fecha':fdate.isoformat(),'rival':str(rival).strip(),'slug':slugify(rival),'cond':str(cond).strip()})
    partidos.sort(key=lambda x:x['fecha'])
    # próximo: el primero con fecha >= hoy
    hoy = datetime.date.today()
    proximo = None
    for p in partidos:
        if datetime.date.fromisoformat(p['fecha']) >= hoy:
            proximo = p; break
    if not proximo and partidos:
        proximo = partidos[-1]  # si ya pasaron todos, el último
    data = {'proximo':proximo,'fixture':partidos}
    with open('proximo_rival.js','w',encoding='utf-8') as f:
        f.write('window.FIXTURE_DATA = ' + json.dumps(data, ensure_ascii=False) + ';\n')
    if proximo:
        print(f"  proximo_rival.js: próximo rival = {proximo['rival']} ({proximo['fecha']})")
    return proximo

def build(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    print(f"Procesando {xlsx_path}...")
    build_videos(wb)
    build_fixture(wb)
    print("LISTO. Subi videos.js y proximo_rival.js al repo.")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python3 build_videos.py videos_casla.xlsx")
        sys.exit(1)
    build(sys.argv[1])
