#!/usr/bin/env python3
# CASLA VOLEY — generar_datos_partidos.py
# Lee CASL_Partidos.xlsx y genera datos_partidos.js
import os, json, re, copy
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl"); exit(1)

BASE  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, "CASL_Partidos.xlsx")
OUT   = os.path.join(BASE, "datos_partidos.js")
SKIP  = {'INDICE', 'GUIA', 'INSTRUCCIONES', 'PRÓXIMO', 'PLANTEL'}

# ── Helpers ──────────────────────────────────────────────────────────
def si(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)))
    except: return 0

def se(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)*100))
    except: return 0

def cn(v):
    if not v: return ''
    s = re.sub(r'\s+',' ', str(v).replace('\xa0',' ')).strip()
    parts = s.split()
    i = 0
    while i < len(parts) and parts[i].isdigit(): i += 1
    return ' '.join(parts[i:]).strip()

def gn(v):
    if not v: return 0
    parts = re.sub(r'\s+',' ', str(v).replace('\xa0',' ')).strip().split()
    return int(parts[0]) if parts and parts[0].isdigit() else 0

def pct(a, b): return round(a*100/b) if b > 0 else 0

def se_pct(v):
    if v is None or str(v).strip() in ('','----'): return 0
    try: return int(round(float(v)*100))
    except: return 0

POS_COLOR = {'CENTRAL':'#f97316','OPUESTO':'#818cf8','PUNTA':'#22c55e',
             'ARMADOR':'#f59e0b','LIBERO':'#06b6d4'}

ORIG = {'X5':4,'V5':4,'X6':2,'V6':2,'X8':9,'V8':9,
        'X1':3,'X7':3,'XM':3,'X2':3,'XP':8,'VP':8,'X4':4}

# ── Leer índice ───────────────────────────────────────────────────────
def leer_indice(wb):
    meta = {}
    if 'INDICE' not in wb.sheetnames: return meta
    ws = wb['INDICE']
    for row in ws.iter_rows(min_row=3, max_row=50, values_only=True):
        if not row[7]: continue
        solapa = str(row[7]).strip()
        # Format fecha - handle both string and datetime
        fecha_raw = row[1]
        if fecha_raw is None:
            fecha_fmt = ''
        elif hasattr(fecha_raw, 'strftime'):
            fecha_fmt = fecha_raw.strftime('%d/%m/%Y')
        else:
            fecha_fmt = str(fecha_raw).strip()
        meta[solapa] = {
            'fecha':      fecha_fmt,
            'rival':      str(row[2]).strip() if row[2] else solapa,
            'torneo':     str(row[3]).strip() if row[3] else '',
            'resultado':  str(row[6]).strip() if row[6] else '',
            'sets_casla': str(row[4]).strip() if row[4] else '',
            'sets_rival': str(row[5]).strip() if row[5] else '',
        }
    return meta

# ── Leer posiciones desde col BM (col 65) ────────────────────────────
# Plantel cargado desde CASLA_PLANTEL
_PLANTEL_CACHE = None

def cargar_plantel():
    global _PLANTEL_CACHE
    if _PLANTEL_CACHE is not None: return _PLANTEL_CACHE
    plantel = {}
    gameplan = os.path.join(BASE, "CASL_GamePlan_4.xlsm")
    if not os.path.exists(gameplan):
        print("  AVISO: No se encontró CASL_GamePlan_4.xlsm — posiciones no disponibles")
        _PLANTEL_CACHE = {}
        return {}
    try:
        wb_gp = load_workbook(gameplan, read_only=True, data_only=True, keep_vba=False)
        if 'CASLA_PLANTEL' in wb_gp.sheetnames:
            ws_pl = wb_gp['CASLA_PLANTEL']
            for row in ws_pl.iter_rows(min_row=2, max_row=25, values_only=True):
                if row[0] and row[1] and row[2]:
                    num = int(row[0]) if str(row[0]).isdigit() else 0
                    nm  = str(row[1]).strip().upper()
                    pos = str(row[2]).strip().upper()
                    plantel[nm] = {'pos': pos, 'num': num}
        print(f"  Plantel: {len(plantel)} jugadores desde CASLA_PLANTEL")
    except Exception as e:
        print(f"  AVISO: Error leyendo plantel: {e}")
    _PLANTEL_CACHE = plantel
    return plantel

def get_pos_num(nm):
    """Get position and number for a player name"""
    plantel = cargar_plantel()
    nm_up = nm.upper()
    # Exact match
    if nm_up in plantel: return plantel[nm_up]['pos'], plantel[nm_up]['num']
    # Partial: last name match
    apellido = nm_up.split()[0] if nm_up else ''
    for k, v in plantel.items():
        if apellido and (apellido in k or k.split()[0] == apellido):
            return v['pos'], v['num']
    return '', 0

def leer_posiciones(ws):
    """Returns empty dict - positions now come from CASLA_PLANTEL via get_pos_num"""
    return {}

# ── Parser ataques ────────────────────────────────────────────────────
def parse_ataques(ws, pos_by_row):
    """
    Structure: player name in col A (rows 4,12), combos in col B
    Blocks of 12 cols per player (A-L, M-X, Y-AJ, AK-AV, AW-BH)
    Zone order: Z1,Z9,Z2,Z6,Z8,Z5,Z7,Z4 (cols C-J), BLOCK OUT col K, TOTAL col L
    """
    all_rows = list(ws.iter_rows(min_row=4, max_row=18, values_only=True))
    attack = {}

    def get_player_from_rows(rows_slice, col_start):
        """Find player name - handles both standard and offset DVW paste formats"""
        # Try col_start and col_start+1 (DVW sometimes pastes offset by 1)
        for offset in [0, 1]:
            cs = col_start + offset
            # Check header row (name in cs, combo header in cs+1)
            nm_cell = rows_slice[0][cs] if cs < len(rows_slice[0]) else None
            bc0 = rows_slice[0][cs+1] if cs+1 < len(rows_slice[0]) else None
            if nm_cell and gn(nm_cell) and (not bc0 or str(bc0).strip() in ('ZONA','ZONA1','')):
                return cn(nm_cell), gn(nm_cell), cs, 1
            # Check inline (name in cs, combo code in cs+1)
            for ri in range(len(rows_slice)):
                row = rows_slice[ri]
                nc = row[cs] if cs < len(row) else None
                bc = row[cs+1] if cs+1 < len(row) else None
                if nc and gn(nc) and bc and str(bc).strip() not in ('','ZONA','ZONA1','TOTALES','BLOCK OUT'):
                    return cn(nc), gn(nc), cs, ri
        return None, 0, col_start, 1

    def parse_block(rows_slice, col_start):
        nm, num, actual_cs, combo_start = get_player_from_rows(rows_slice, col_start)
        if not nm: return
        if nm not in attack:
            attack[nm] = {'num':num, 'combos':{}}

        for ri in range(combo_start, len(rows_slice)):
            row = rows_slice[ri]
            nc = row[actual_cs] if actual_cs < len(row) else None
            if nc and isinstance(nc, str) and 'OUTSIDE' in nc.upper(): break
            if nc and gn(nc) > 0 and ri > combo_start:
                bc_check = row[actual_cs+1] if actual_cs+1 < len(row) else None
                if not bc_check or str(bc_check).strip() in ('','ZONA','ZONA1'): break

            bc = row[actual_cs+1] if actual_cs+1 < len(row) else None
            if not bc or str(bc).strip() in ('','ZONA','ZONA1','TOTALES','COMBO'): continue
            cod = str(bc).strip()
            # Total is always 11 cols after actual_cs
            tc = row[actual_cs+11] if actual_cs+11 < len(row) else None
            tot = si(tc)
            zones = {}
            for di, zn in enumerate([1,9,2,6,8,5,7,4]):
                ci = actual_cs+2+di
                zones[zn] = si(row[ci] if ci < len(row) else None)
            if tot == 0: tot = sum(zones.values())
            if cod and (tot > 0 or any(v > 0 for v in zones.values())):
                attack[nm]['combos'][cod] = {'total':tot, 'zones':zones}

    for col_start in [0, 12, 24, 36, 48]:
        parse_block(all_rows[0:8],  col_start)
        parse_block(all_rows[8:15], col_start)

    return attack

# ── Parser saques ─────────────────────────────────────────────────────
def parse_saques(ws):
    """
    Structure rows 22-34:
    Col A=nombre, Col B=zona (Z1/Z6/Z5), Cols C-H=destinos, Col I=total
    Blocks of 9 cols. Groups: rows 22-27 (SM Z1,Z6,Z5 + SQ Z1,Z6,Z5), rows 29-34 (same)
    """
    saq_rows = list(ws.iter_rows(min_row=21, max_row=34, values_only=True))
    serve = {}

    def parse_group(rows_slice):
        # Find players: when col A has a name with number
        for block in range(10):
            col = block * 9
            if col >= len(rows_slice[0] or []): break
            # Row 0 of group = first player Z1 SM
            nm_cell = rows_slice[0][col] if col < len(rows_slice[0]) else None
            if not nm_cell or not gn(nm_cell): continue
            nm = cn(nm_cell)
            num = gn(nm_cell)
            if nm not in serve: serve[nm] = {'num':num, 'sm':{}, 'sq':{}}

            # SM: rows 0-2 (Z1,Z6,Z5)
            # SQ: rows 3-5 (Z1,Z6,Z5)
            dest_sm = {}; dest_sq = {}
            for ri in range(3):
                row = rows_slice[ri] if ri < len(rows_slice) else []
                for di, dz in enumerate([1,9,6,8,5,7]):
                    ci = col+2+di
                    val = si(row[ci] if ci < len(row) else None)
                    if val > 0: dest_sm[dz] = dest_sm.get(dz,0) + val
            for ri in range(3,6):
                row = rows_slice[ri] if ri < len(rows_slice) else []
                for di, dz in enumerate([1,9,6,8,5,7]):
                    ci = col+2+di
                    val = si(row[ci] if ci < len(row) else None)
                    if val > 0: dest_sq[dz] = dest_sq.get(dz,0) + val

            serve[nm]['sm'] = dest_sm
            serve[nm]['sq'] = dest_sq

    # Group 1: rows 22-27 (index 1-6, skipping header row 21)
    parse_group(saq_rows[1:7])
    # Group 2: rows 29-34 (index 8-13)
    parse_group(saq_rows[8:14])
    # Also parse col J block (second player in each group, 9 cols after col A)
    def parse_group_j(rows_slice):
        # Col J = col index 9, blocks start at col 9, 18, 27...
        for block in range(10):
            col = block * 9 + 9  # offset by 9 for col J start
            if col >= len(rows_slice[0] or []): break
            nm_cell = rows_slice[0][col] if col < len(rows_slice[0]) else None
            if not nm_cell or not gn(nm_cell): continue
            nm = cn(nm_cell); num = gn(nm_cell)
            if nm not in serve: serve[nm] = {'num':num, 'sm':{}, 'sq':{}}
            dest_sm = {}; dest_sq = {}
            for ri in range(3):
                row = rows_slice[ri] if ri < len(rows_slice) else []
                for di, dz in enumerate([1,9,6,8,5,7]):
                    ci = col+2+di
                    val = si(row[ci] if ci < len(row) else None)
                    if val > 0: dest_sm[dz] = dest_sm.get(dz,0) + val
            for ri in range(3,6):
                row = rows_slice[ri] if ri < len(rows_slice) else []
                for di, dz in enumerate([1,9,6,8,5,7]):
                    ci = col+2+di
                    val = si(row[ci] if ci < len(row) else None)
                    if val > 0: dest_sq[dz] = dest_sq.get(dz,0) + val
            serve[nm]['sm'] = dest_sm; serve[nm]['sq'] = dest_sq
    parse_group_j(saq_rows[1:7])
    parse_group_j(saq_rows[8:14])

    return serve


# ── Parser de recepción (rows 89-125, mismo formato que CASLA) ───────
def parse_recepcion_partido(ws):
    rec_all = list(ws.iter_rows(min_row=89, max_row=126, values_only=True))
    if len(rec_all) < 2: return {}
    rec_header = rec_all[1]
    rec_cols = []
    for ci, v in enumerate(rec_header):
        if v and isinstance(v, str) and any(c.isdigit() for c in str(v)[:3]):
            rec_cols.append((ci, cn(v), gn(v)))

    recepcion = {}
    def get_rec(row_idx, col_start):
        if row_idx >= len(rec_all): return {'tot':0,'eff':0,'pos':0,'neg':0}
        row = rec_all[row_idx]
        tot      = si(row[col_start+2] if col_start+2 < len(row) else None)
        eff      = se_pct(row[col_start+3] if col_start+3 < len(row) else None)
        cnt_perf = si(row[col_start+4] if col_start+4 < len(row) else None)
        cnt_pos  = si(row[col_start+5] if col_start+5 < len(row) else None)
        cnt_neg  = si(row[col_start+6] if col_start+6 < len(row) else None)
        cnt_err  = si(row[col_start+7] if col_start+7 < len(row) else None)
        return {'tot':tot,'eff':eff,
                'pos':pct(cnt_perf+cnt_pos,tot),'neg':pct(cnt_neg+cnt_err,tot)}

    for col_start, name, num in rec_cols:
        flotado = {
            'desde_z1':{'total':get_rec(11,col_start),'p1':get_rec(12,col_start),'p6':get_rec(13,col_start),'p5':get_rec(14,col_start)},
            'desde_z6':{'total':get_rec(15,col_start),'p1':get_rec(16,col_start),'p6':get_rec(17,col_start),'p5':get_rec(18,col_start)},
            'desde_z5':{'total':get_rec(19,col_start),'p1':get_rec(20,col_start),'p6':get_rec(21,col_start),'p5':get_rec(22,col_start)},
        }
        potencia = {
            'desde_z1':{'total':get_rec(25,col_start),'p1':get_rec(26,col_start),'p6':get_rec(27,col_start),'p5':get_rec(28,col_start)},
            'desde_z6':{'total':get_rec(29,col_start),'p1':get_rec(30,col_start),'p6':get_rec(31,col_start),'p5':get_rec(32,col_start)},
            'desde_z5':{'total':get_rec(33,col_start),'p1':get_rec(34,col_start),'p6':get_rec(35,col_start),'p5':get_rec(36,col_start)},
        }
        recepcion[name] = {'num':num,'flotado':flotado,'potencia':potencia}
    return recepcion

# ── Parser de armadores (rows 67-86) ────────────────────────────────
def parse_armadores_partido(ws):
    # All rows as list (row67=idx0, row68=idx1, row69=idx2, ...)
    # Titular:  header=idx1(row68), data P1-P6=idx2-7(row69-74), SD/TR=idx8-9(row75-76)
    # Suplente: header=idx11(row78), data P1-P6=idx12-17(row79-84), SD/TR=idx18-19(row85-86)
    # SO/TR grid: idx14-17 (rows 81-84) for titular, idx24-27 (rows 91-94) for suplente
    # Cols (0-indexed): AD=29,AE=30,AF=31,AG=32,AH=33,AI=34 | AB=27,AC=28

    rows = list(ws.iter_rows(min_row=67, max_row=95, values_only=True))

    def sp(v):
        if v is None or str(v).strip() in ('','----','None'): return 0
        try: return int(round(float(v)*100))
        except: return 0

    def read_arm(hdr_idx, grid_base_idx):
        hdr = rows[hdr_idx] if hdr_idx < len(rows) else []
        nombre = cn(hdr[29]) if len(hdr)>29 and hdr[29] else ''

        pos_labels = ['P1','P6','P5','P4','P3','P2']
        pills = []
        for i in range(6):
            ri = hdr_idx + 1 + i
            r = rows[ri] if ri < len(rows) else []
            lbl   = str(r[29]).strip() if len(r)>29 and r[29] else pos_labels[i]
            eff   = sp(r[30]) if len(r)>30 else 0   # AE = EFF%
            tot   = si(r[31]) if len(r)>31 else 0   # AF = TOT
            pts   = si(r[32]) if len(r)>32 else 0   # AG = #
            pct_v = sp(r[33]) if len(r)>33 else 0   # AH = #%
            err_v = sp(r[34]) if len(r)>34 else 0   # AI = /=%
            pills.append({'label':lbl,'eff':eff,'tot':tot,'pts':pts,'pts_pct':pct_v,'err_pct':err_v})

        # SO and TR pills (idx hdr+7, hdr+8)
        for off, lbl in [(7,'SO'),(8,'TR')]:
            ri = hdr_idx + off
            r = rows[ri] if ri < len(rows) else []
            eff   = sp(r[30]) if len(r)>30 else 0
            tot   = si(r[31]) if len(r)>31 else 0
            pts   = si(r[32]) if len(r)>32 else 0
            pct_v = sp(r[33]) if len(r)>33 else 0
            err_v = sp(r[34]) if len(r)>34 else 0
            pills.append({'label':lbl,'eff':eff,'tot':tot,'pts':pts,'pts_pct':pct_v,'err_pct':err_v})

        # Rec pills: AB(27)=EFF% for each rotation row
        rec_labels = ['SO','REC #','REC +','REC !','REC -','TRANS']
        extra_pills = []
        for i in range(6):
            ri = hdr_idx + 1 + i
            r = rows[ri] if ri < len(rows) else []
            eff = sp(r[27]) if len(r)>27 else 0   # AB = EFF%
            extra_pills.append({'label':rec_labels[i],'eff':eff})

        # SO/TR distribution grid
        # rows grid_base_idx to grid_base_idx+3
        # row+0 (odd): A=soP4tot,C=soP3tot,E=soP2tot, G=trP4tot,I=trP3tot,K=trP2tot
        # row+1 (even): A=soP4pts,B=soP4pct,C=soP3pts,D=soP3pct... G=trP4pts...
        # row+2 (odd): C=soP6tot,E=soP1tot, I=trP6tot,K=trP1tot
        # row+3 (even): C=soP6pts,D=soP6pct,E=soP1pts,F=soP1pct, I=trP6pts...
        g = grid_base_idx
        r0 = rows[g]   if g   < len(rows) else []
        r1 = rows[g+1] if g+1 < len(rows) else []
        r2 = rows[g+2] if g+2 < len(rows) else []
        r3 = rows[g+3] if g+3 < len(rows) else []

        def gv(row, idx): return si(row[idx]) if idx < len(row) else 0
        def gp(row, idx): return sp(row[idx]) if idx < len(row) else 0

        so_dist = [
            {'pos':'P4','tot':gv(r0,0),'pts':gv(r1,0),'pct':gp(r1,1)},
            {'pos':'P3','tot':gv(r0,2),'pts':gv(r1,2),'pct':gp(r1,3)},
            {'pos':'P2','tot':gv(r0,4),'pts':gv(r1,4),'pct':gp(r1,5)},
            {'pos':'P1','tot':gv(r2,4),'pts':gv(r3,4),'pct':gp(r3,5)},
            {'pos':'P6','tot':gv(r2,2),'pts':gv(r3,2),'pct':gp(r3,3)},
        ]
        tr_dist = [
            {'pos':'P4','tot':gv(r0,6),'pts':gv(r1,6),'pct':gp(r1,7)},
            {'pos':'P3','tot':gv(r0,8),'pts':gv(r1,8),'pct':gp(r1,9)},
            {'pos':'P2','tot':gv(r0,10),'pts':gv(r1,10),'pct':gp(r1,11)},
            {'pos':'P1','tot':gv(r2,10),'pts':gv(r3,10),'pct':gp(r3,11)},
            {'pos':'P6','tot':gv(r2,8),'pts':gv(r3,8),'pct':gp(r3,9)},
        ]

        # Build rotaciones from dist data
        pos_rot = ['P1','P6','P5','P4','P3','P2']
        col_off = 0 if hdr_idx == 1 else 13
        rots = []
        for i in range(6):
            r_tot = rows[1 + i*2] if 1+i*2 < len(rows) else []
            r_pts = rows[2 + i*2] if 2+i*2 < len(rows) else []
            c = col_off
            tot2 = si(r_tot[10+c] if 10+c<len(r_tot) else None)
            dZ4  = si(r_tot[2+c]  if 2+c <len(r_tot) else None)
            dZ3  = si(r_tot[4+c]  if 4+c <len(r_tot) else None)
            dZ2  = si(r_tot[6+c]  if 6+c <len(r_tot) else None)
            dZ6  = si(r_tot[8+c]  if 8+c <len(r_tot) else None)
            pZ4  = si(r_pts[2+c]  if 2+c <len(r_pts) else None)
            pZ3  = si(r_pts[4+c]  if 4+c <len(r_pts) else None)
            pZ2  = si(r_pts[6+c]  if 6+c <len(r_pts) else None)
            pZ6  = si(r_pts[8+c]  if 8+c <len(r_pts) else None)
            rots.append({'pos':pos_rot[i],'total':tot2,'dist':[
                {'zona':4,'tot':dZ4,'pts':pZ4,'pct':pct(dZ4,tot2),'pct_p':pct(pZ4,dZ4)},
                {'zona':3,'tot':dZ3,'pts':pZ3,'pct':pct(dZ3,tot2),'pct_p':pct(pZ3,dZ3)},
                {'zona':2,'tot':dZ2,'pts':pZ2,'pct':pct(dZ2,tot2),'pct_p':pct(pZ2,dZ2)},
                {'zona':6,'tot':dZ6,'pts':pZ6,'pct':pct(dZ6,tot2),'pct_p':pct(pZ6,dZ6)},
            ]})

        return {'nombre':nombre,'rotaciones':rots,'pills':pills,
                'extra_pills':extra_pills,
                'so_dist':so_dist,'tr_dist':tr_dist,
                'recepcion':[],'so':{},'tr':{}}

    tit = read_arm(1,  14)   # header=row68, grid=rows81-84
    sup = read_arm(11, 24)   # header=row78, grid=rows91-94

    # Override nombre from row 67
    hdr0 = rows[0]
    if len(hdr0)>1  and hdr0[1]:  tit['nombre'] = cn(hdr0[1])
    if len(hdr0)>14 and hdr0[14]: sup['nombre'] = cn(hdr0[14])

    return {'titular': tit, 'suplente': sup}


# ── Parser de objetivos/baterías (rows 128-144) ──────────────────────
def parse_objetivos(ws):
    # Row 128 = header, row 129 = equipo, rows 130-144 = jugadores
    # Cols (0-indexed):
    #   G(6)=%SQ, M(12)=%REC, P(15)=%BLQ#+, T(19)=%PTBLQ,
    #   Y(24)=%ATQCENTRAL, AD(29)=%ATQALTA, AI(34)=%ATQRAPIDA,
    #   AN(39)=%ATQRP, AS(44)=%ATQRI, AX(49)=%ATQRM, BC(54)=%ATQTR
    COL_MAP = {
        'sq':6, 'rec':12, 'bqpos':15, 'bqpt':19,
        'atqq':24, 'atqhb':29, 'atqx':34,
        'atqrp':39, 'atqri':44, 'atqrm':49, 'atqtr':54
    }
    obj_rows = list(ws.iter_rows(min_row=128, max_row=144, values_only=True))
    if len(obj_rows) < 2: return {}

    result = {}

    def read_row(row):
        vals = {}
        for key, cidx in COL_MAP.items():
            v = row[cidx] if cidx < len(row) else None
            if v is not None and str(v).strip() not in ('','----'):
                try: vals[key] = int(round(float(v) * 100))
                except: vals[key] = None
            else:
                vals[key] = None
        return vals

    # Row 1 = equipo (idx 1 = row 129)
    if len(obj_rows) > 1:
        result['__equipo__'] = read_row(obj_rows[1])

    # Rows 2-16 = jugadores (idx 2-16 = rows 130-144)
    for ri in range(2, len(obj_rows)):
        row = obj_rows[ri]
        nm_cell = row[0] if row[0] else None
        if not nm_cell: continue
        nm = cn(nm_cell)
        if nm:
            result[nm] = read_row(row)

    return result

# ── Leer partido completo ─────────────────────────────────────────────
def leer_partido(ws, meta):
    pos_by_row = leer_posiciones(ws)
    attack = parse_ataques(ws, pos_by_row)
    serve  = parse_saques(ws)

    player_pos = leer_posiciones(ws)

    # Parse stats (rows 36-64)
    # SM section: row 36=header, rows 37-49=players
    # SQ section: row 51=header, rows 52-64=players
    # Attack combos by column (0-indexed):
    # SM row cols: A=nombre(0),B=eff(1),C=tot(2),D=pts(3),E=plus(4)
    #   X8:  H(7)=nom, I(8)=eff, J(9)=tot, K(10)=pts, L(11)=%pts
    #   X5:  O(14)=nom,P(15)=eff,Q(16)=tot,R(17)=pts,S(18)=%pts
    #   X6:  V(21)=nom,W(22)=eff,X(23)=tot,Y(24)=pts,Z(25)=%pts
    #   X1:  AC(28)=nom,AD(29)=eff,AE(30)=tot,AF(31)=pts,AG(32)=%pts
    #   X7:  AJ(35)=nom,AK(36)=eff,AL(37)=tot,AM(38)=pts,AN(39)=%pts
    # SQ row cols same structure but:
    #   V8:H, V5:O, V6:V, XM:AC, X2:AJ
    # Attack combo columns (0-indexed): nom, eff, tot, pts, pct_pts
    # SM row: X8(7), X5(14), X6(21), X1(28), X7(35), XP(42)
    # SQ row: V8(7), V5(14), V6(21), XM(28), X2(35), VP(42)
    ATK_COLS = {
        'sm': [('X8',7),('X5',14),('X6',21),('X1',28),('X7',35),('XP',42)],
        'sq': [('V8',7),('V5',14),('V6',21),('XM',28),('X2',35),('VP',42)],
    }
    # XP is a variant of X5/V5 - not separately tracked in stats

    stats = {}
    stat_rows = list(ws.iter_rows(min_row=36, max_row=65, values_only=True))

    def read_stat_section(rows, tipo):
        for row in rows:
            nm_a = row[0] if row[0] else None
            if not nm_a or not gn(nm_a): continue
            nm = cn(nm_a)
            if nm not in stats:
                stats[nm] = {'sm_eff':0,'sm_tot':0,'sm_pts':0,'sm_plus':0,
                             'sq_eff':0,'sq_tot':0,'sq_pts':0,'sq_plus':0,'atk':{}}
            # SM/SQ overall
            eff  = se(row[1]) if len(row)>1 else 0
            tot  = si(row[2]) if len(row)>2 else 0
            pts  = si(row[3]) if len(row)>3 else 0
            plus = si(row[4]) if len(row)>4 else 0
            stats[nm][tipo+'_eff']  = eff
            stats[nm][tipo+'_tot']  = tot
            stats[nm][tipo+'_pts']  = pts
            stats[nm][tipo+'_plus'] = plus
            # Attack combos
            for cod, base in ATK_COLS[tipo.lower()]:
                nm_b = row[base] if base < len(row) else None
                if not nm_b or not gn(nm_b): continue
                a_eff = se(row[base+1]) if base+1 < len(row) else 0
                a_tot = si(row[base+2]) if base+2 < len(row) else 0
                a_pts = si(row[base+3]) if base+3 < len(row) else 0
                a_pct = se(row[base+4]) if base+4 < len(row) else 0
                if a_tot > 0 or a_eff != 0:
                    stats[nm]['atk'][cod] = {
                        'eff': a_eff, 'tot': a_tot,
                        'pts': a_pts, 'pts_pct': a_pct
                    }

    # SM section: rows 37-49 (indices 1-13)
    read_stat_section(stat_rows[1:14], 'sm')
    # SQ section: rows 52-64 (indices 16-28)
    read_stat_section(stat_rows[16:29], 'sq')

    recepcion = parse_recepcion_partido(ws)
    armadores = parse_armadores_partido(ws)
    objetivos = parse_objetivos(ws)

    return {
        'rival':      meta.get('rival',''),
        'fecha':      meta.get('fecha',''),
        'torneo':     meta.get('torneo',''),
        'resultado':  meta.get('resultado',''),
        'sets_casla': meta.get('sets_casla',''),
        'sets_rival': meta.get('sets_rival',''),
        'attack':     attack,
        'serve':      serve,
        'player_pos': player_pos,
        'recepcion':  recepcion,
        'armadores':  armadores,
        'objetivos':  objetivos,
        'stats':      stats,
    }

# ── Acumular todos los partidos ───────────────────────────────────────
def acumular(partidos_data):
    all_players = {}

    for pd in partidos_data:
        pp = pd['player_pos']

        for nm, data in pd['attack'].items():
            pos_pl, num_pl = get_pos_num(nm)
            pos = pos_pl or pp.get(nm, '')
            num = num_pl or data['num']
            if nm not in all_players:
                all_players[nm] = {'num':num,'pos':pos,'combos':{},'serve_sm':{},'serve_sq':{}}
            if pos and not all_players[nm]['pos']:
                all_players[nm]['pos'] = pos
            if num and not all_players[nm]['num']:
                all_players[nm]['num'] = num
            for cod, a in data['combos'].items():
                if cod not in all_players[nm]['combos']:
                    all_players[nm]['combos'][cod] = {'total':0,'zones':{}}
                all_players[nm]['combos'][cod]['total'] += a['total']
                for z,v in a['zones'].items():
                    all_players[nm]['combos'][cod]['zones'][z] =                         all_players[nm]['combos'][cod]['zones'].get(z,0) + v

            # stats accumulated separately below

        for nm, data in pd['serve'].items():
            pos_pl, num_pl = get_pos_num(nm)
            pos = pos_pl or pp.get(nm, '')
            num_srv = num_pl or data.get('num', 0)
            if nm not in all_players:
                all_players[nm] = {'num':num_srv,'pos':pos,'combos':{},'serve_sm':{},'serve_sq':{}}
            else:
                if pos and not all_players[nm]['pos']: all_players[nm]['pos'] = pos
                if num_srv and not all_players[nm]['num']: all_players[nm]['num'] = num_srv
            if pos and not all_players[nm]['pos']:
                all_players[nm]['pos'] = pos
            for dz,v in data.get('sm',{}).items():
                all_players[nm]['serve_sm'][dz] = all_players[nm]['serve_sm'].get(dz,0)+v
            for dz,v in data.get('sq',{}).items():
                all_players[nm]['serve_sq'][dz] = all_players[nm]['serve_sq'].get(dz,0)+v
            # stats accumulated below

        # Recepcion: store per jugador (take latest non-empty)
        for nm, rec in pd.get('recepcion',{}).items():
            if nm not in all_players:
                # Add player even if not in attack/serve
                pos_pl, num_pl = get_pos_num(nm)
                all_players[nm] = {'num':num_pl,'pos':pos_pl,'combos':{},
                                   'serve_sm':{},'serve_sq':{}}
            if not all_players[nm].get('recepcion'):
                all_players[nm]['recepcion'] = rec

        # Objetivos: weighted average across partidos
        for nm, obj in pd.get('objetivos',{}).items():
            if nm == '__equipo__': continue
            # Create player if not exists (some players only appear in objetivos)
            if nm not in all_players:
                pos_pl, num_pl = get_pos_num(nm)
                all_players[nm] = {'num':num_pl,'pos':pos_pl,'combos':{},
                                   'serve_sm':{},'serve_sq':{}}
            if 'objetivos' not in all_players[nm]:
                all_players[nm]['objetivos'] = {k:{'sum':0,'n':0} for k in obj}
            for k, v in obj.items():
                if v is not None:
                    if k not in all_players[nm]['objetivos']:
                        all_players[nm]['objetivos'][k] = {'sum':0,'n':0}
                    all_players[nm]['objetivos'][k]['sum'] += v
                    all_players[nm]['objetivos'][k]['n']   += 1

        # Equipo objetivos
        eq_obj = pd.get('objetivos',{}).get('__equipo__',{})
        if eq_obj:
            if '__equipo__' not in all_players:
                all_players['__equipo__'] = {'num':0,'pos':'','combos':{},'serve_sm':{},'serve_sq':{}}
            if 'objetivos' not in all_players['__equipo__']:
                all_players['__equipo__']['objetivos'] = {k:{'sum':0,'n':0} for k in eq_obj}
            for k, v in eq_obj.items():
                if v is not None:
                    all_players['__equipo__']['objetivos'][k]['sum'] = all_players['__equipo__']['objetivos'][k].get('sum',0) + v
                    all_players['__equipo__']['objetivos'][k]['n']   = all_players['__equipo__']['objetivos'][k].get('n',0) + 1

    # Armadores: accumulate as acumulado
    all_armadores = {}
    for pd in partidos_data:
        arm = pd.get('armadores', {})
        for tipo in ['titular','suplente']:
            a = arm.get(tipo,{})
            if not a or not a.get('nombre'): continue
            if tipo not in all_armadores:
                all_armadores[tipo] = {'nombre':a['nombre'],'rotaciones':[],'pills':a.get('pills',[]),
                                        'recepcion':[],'so':{},'tr':{}}
            # Accumulate rotaciones
            for ri, rot in enumerate(a.get('rotaciones',[])):
                if ri >= len(all_armadores[tipo].get('rotaciones',[])):
                    all_armadores[tipo]['rotaciones'].append({'pos':rot['pos'],'total':0,'dist':rot['dist'][:] })
                all_armadores[tipo]['rotaciones'][ri]['total'] += rot['total']

    # ── Accumulate stats once per player per partido ─────────────────
    already_done = set()
    for pd in partidos_data:
        for nm in list(pd.get('stats', {}).keys()):
            key = (nm, pd.get('rival',''))
            if key in already_done: continue
            already_done.add(key)
            if nm not in all_players: continue
            pd_stats = pd['stats'][nm]
            if 'stats' not in all_players[nm]:
                all_players[nm]['stats'] = {'sm_eff':0,'sm_tot':0,'sm_pts':0,'sm_plus':0,
                                             'sq_eff':0,'sq_tot':0,'sq_pts':0,'sq_plus':0,'atk':{}}
            s = all_players[nm]['stats']
            for tipo in ['sm','sq']:
                new_tot = pd_stats.get(tipo+'_tot',0)
                new_eff = pd_stats.get(tipo+'_eff',0)
                old_tot = s.get(tipo+'_tot',0)
                old_eff = s.get(tipo+'_eff',0)
                combined = old_tot + new_tot
                s[tipo+'_eff']  = round((old_eff*old_tot + new_eff*new_tot)/combined) if combined else 0
                s[tipo+'_tot']  = combined
                s[tipo+'_pts']  = s.get(tipo+'_pts',0)  + pd_stats.get(tipo+'_pts',0)
                s[tipo+'_plus'] = s.get(tipo+'_plus',0) + pd_stats.get(tipo+'_plus',0)
            for cod2, av in pd_stats.get('atk',{}).items():
                if cod2 not in s['atk']:
                    s['atk'][cod2] = {'eff':0,'tot':0,'pts':0,'pts_pct':0}
                new_tot = av.get('tot',0)
                new_eff = av.get('eff',0) or 0
                old_tot = s['atk'][cod2]['tot']
                old_eff = s['atk'][cod2]['eff']
                combined = old_tot + new_tot
                # Weighted average EFF
                s['atk'][cod2]['eff'] = round((old_eff*old_tot + new_eff*new_tot)/combined) if combined else 0
                s['atk'][cod2]['tot'] = combined
                s['atk'][cod2]['pts'] += av.get('pts',0)
                # pts_pct: weighted average of #% from Excel
                old_pct = s['atk'][cod2].get('pts_pct',0)
                new_pct = av.get('pts_pct',0)
                s['atk'][cod2]['pts_pct'] = round((old_pct*old_tot + new_pct*new_tot)/combined) if combined else 0

    return all_players, all_armadores

# ── Build jugadores array ─────────────────────────────────────────────
def build_jugadores(all_players, stats_override=None):
    jugadores = []
    for nm in sorted((k for k in all_players if k != '__equipo__'), key=lambda n: all_players[n]['num']):
        info  = all_players[nm]
        pos   = info['pos']
        color = POS_COLOR.get(pos,'#64748b')
        num   = info['num']

        # Ataques
        ataques = []
        # Use per-partido stats if provided, else accumulated
        st = stats_override.get(nm, info.get('stats', {})) if stats_override else info.get('stats', {})
        for cod, a in info['combos'].items():
            tot   = a['total'] or sum(a['zones'].values())
            zones = a['zones']
            orig  = ORIG.get(cod,3)
            dest_sorted = sorted([(z,v) for z,v in zones.items() if v>0], key=lambda x:-x[1])
            destinos = [{'z':z,'pct':pct(v,tot)} for z,v in dest_sorted[:4]]
            # EFF from stats (X8/V8 only), otherwise null
            atk_st = st.get('atk', {}).get(cod, {})
            eff_raw = atk_st.get('eff', None) if atk_st else None
            eff = eff_raw if eff_raw is not None else None
            pts = atk_st.get('pts', 0) if atk_st else 0
            st_tot = atk_st.get('tot', 0) if atk_st else 0
            use_tot = st_tot if st_tot else tot
            # Use #% directly from Excel (weighted avg across partidos)
            pts_pct = atk_st.get('pts_pct', 0) if atk_st else 0
            if not pts_pct and pts and use_tot:
                pts_pct = pct(pts, use_tot)
            ataques.append({'cod':cod,'orig':orig,'destinos':destinos,
                            'eff':eff,'tot':use_tot,'pts':pts,'slash':0,'err':0,
                            'video':None,'pts_pct':pts_pct})

        # Saques
        saques = []
        for cod, tipo_label, dest_data, st_key in [
            ('SM','FLOTADO', info['serve_sm'], 'sm'),
            ('SQ','POTENCIA',info['serve_sq'], 'sq')]:
            tot_dest = sum(dest_data.values()) or 0
            destinos = [{'z':z,'pct':pct(v,tot_dest)}
                        for z,v in sorted(dest_data.items(),key=lambda x:-x[1])[:6]] if tot_dest else []
            # EFF and PTS from stats
            s_eff  = st.get(st_key+'_eff', 0)
            s_tot  = st.get(st_key+'_tot', 0) or tot_dest
            s_pts  = st.get(st_key+'_pts', 0)
            s_plus = st.get(st_key+'_plus', 0)
            s_pts_pct = pct(s_pts, s_tot) if s_tot else 0
            saques.append({'cod':cod,'tipo':tipo_label,'orig':6,'destinos':destinos,
                           'eff':s_eff,'tot':s_tot,'pts':s_pts,
                           'plus':s_plus,'slash':0,'err':0,'video':None,
                           'pts_pct':s_pts_pct})

        j = {'num':num,'nombre':f"{num} {nm}",'pos':pos,'color':color,
             'info':{},'ataques':ataques,'saques':saques,'video':0,'recepcion':info.get('recepcion',{})}
        jugadores.append(j)

    return jugadores

# ── Main ──────────────────────────────────────────────────────────────
def main():
    print("="*50)
    print("CASLA VOLEY — Generar datos_partidos.js")
    print("="*50)

    if not os.path.exists(EXCEL):
        print(f"ERROR: No se encontró {EXCEL}"); exit(1)

    wb = load_workbook(EXCEL, read_only=True, data_only=True, keep_vba=False)
    print(f"Hojas: {wb.sheetnames}\n")

    meta_idx = leer_indice(wb)

    # Match sheet names to indice (fuzzy match)
    def get_meta(sheet_name):
        if sheet_name in meta_idx: return meta_idx[sheet_name]
        # Try fuzzy: sheet CIUDAD matches CIUDAD VOLEY
        for k, v in meta_idx.items():
            if sheet_name in k or k in sheet_name:
                return v
        return {'rival':sheet_name,'fecha':'','torneo':'','resultado':'',
                'sets_casla':'','sets_rival':''}

    partidos_procesados = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP: continue
        ws = wb[sheet_name]
        meta = get_meta(sheet_name)
        print(f"Procesando: {sheet_name} ({meta['fecha']} vs {meta['rival']})")
        pd_data = leer_partido(ws, meta)
        partidos_procesados.append({'nombre':sheet_name, **pd_data})

    # Acumulate
    all_players, all_armadores = acumular(partidos_procesados)
    jugadores   = build_jugadores(all_players)

    print(f"\n✓ {len(jugadores)} jugadores:")
    for j in jugadores:
        nm = re.sub(r'^\d+\s*','',j['nombre'])
        na = len([a for a in j['ataques'] if a['tot']>0])
        ns = sum(s['tot'] for s in j['saques'])
        print(f"  #{j['num']} {nm} ({j['pos']}) — {na} combos, {ns} saques")

    # Per partido
    partidos_list = []
    for pd in partidos_procesados:
        import copy
        pd_copy = dict(pd)
        pd_copy['stats'] = json.loads(json.dumps(pd.get('stats', {})))
        pp, pp_arm = acumular([pd_copy])
        jj = build_jugadores(pp, stats_override=pd.get('stats', {}))
        partidos_list.append({
            'nombre':     pd['nombre'],
            'rival':      pd['rival'],
            'fecha':      pd['fecha'],
            'torneo':     pd['torneo'],
            'resultado':  pd['resultado'],
            'sets_casla': pd['sets_casla'],
            'sets_rival': pd['sets_rival'],
            'jugadores':  jj,
            'armadores':  json.loads(json.dumps(pd.get('armadores', {}),default=str)),
            'objetivos':  pd.get('objetivos', {}),
            'equipo_obj': pd.get('objetivos', {}).get('__equipo__', {}),
        })

    generado = datetime.now().strftime('%d/%m/%Y %H:%M')
    meta_list = [
        {'nombre':p['nombre'],'rival':p['rival'],'fecha':p['fecha'],
         'torneo':p['torneo'],'resultado':p['resultado'],
         'sets_casla':p['sets_casla'],'sets_rival':p['sets_rival']}
        for p in partidos_procesados
    ]
    eq_raw = all_players.get('__equipo__', {}).get('objetivos', {})
    equipo_obj = {k: (round(v['sum']/v['n']) if v.get('n') else None)
                  for k, v in eq_raw.items()}
    js = f"""// datos_partidos.js — {generado}
const PARTIDOS_GENERADO = "{generado}";
const PARTIDOS_TOTAL = {len(partidos_procesados)};
const PARTIDOS_META = {json.dumps(meta_list, ensure_ascii=False, indent=2)};
const PARTIDOS_JUGADORES = {json.dumps(jugadores, ensure_ascii=False, indent=2)};

const PARTIDOS_ARMADOR = {json.dumps(all_armadores, ensure_ascii=False, indent=2)};
const PARTIDOS_INDIVIDUAL = {json.dumps(partidos_list, ensure_ascii=False, indent=2)};

// Objetivos del equipo acumulado
const PARTIDOS_EQUIPO_OBJ = {json.dumps(equipo_obj, ensure_ascii=False)};
"""
    with open(OUT,'w',encoding='utf-8') as f: f.write(js)
    print(f"\n✓ datos_partidos.js generado ({len(js)//1024}KB) — subir a GitHub")

if __name__ == '__main__':
    main()
